import requests
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook
wb = Workbook()


def cf_decode_email(some_soup):
    email_raw = some_soup.findAll('a')[0]['data-cfemail']
    r = int(email_raw[:2], 16)
    email = ''.join([chr(int(email_raw[l:l+2], 16) ^ r) for l in range(2, len(email_raw), 2)])
    return email


def login(username, password):
    s = requests.session()
    resp = s.post('https://joffice.jeunesseglobal.com/login.asp')
    data_sec = {
        'Username': username,
        'password': password,
    }
    resp_sec = s.post('https://security-api.jeunesseglobal.com/api/v2/token', data=data_sec)
    token = resp_sec.content.decode('utf-8', 'strict')
    s.cookies.__setitem__('token', token)
    data_log = {
        'Username': username,
        'pw': password,
    }
    resp_log = s.post('https://joffice.jeunesseglobal.com/login.asp?ReturnURL=', data=data_log)

    resp_orders = s.get(
        'https://cart-api.jeunesseglobal.com/api/v1/orders?mainId=3014062&orderNumber=&startDate=&endDate=&paidStatus=&orderType=&offset=0&limit=10&orderBy=-mainOrdersPK&enrollerSiteUrl=&mainSiteUrl=')
    orders_dict = json.loads(resp_orders.content.decode())
    orders_list = orders_dict['orders']
    print(orders_list)
    return orders_list, s


def letter_transform(order):
    order_number = order['mainOrdersPK']

    order_url = 'https://joffice.jeunesseglobal.com/SingleInvoice.asp?uordernumber=' + str(order_number)
    resp_order = s.get(order_url)
    soup = BeautifulSoup(resp_order.text, 'html.parser')

    email_decoded = cf_decode_email(soup)

    print(email_decoded)
    print(type(email_decoded))
    print('------------------------------------------------------------------------------')

    tr = soup.findAll('tr')

    column_1 = []
    column_2 = []
    column_3 = []
    column_4 = []
    column_5 = []
    column_6 = []
    list_of_columns = [column_1, column_2, column_3, column_4, column_5, column_6]

    for i in tr:

        td = i.get_text()
        td_list = [x for x in td.split('\n') if x != '']
        for j in list_of_columns:
            try:
                j.append(td_list[list_of_columns.index(j)])
            except IndexError:
                j.append('')

    ws = wb.create_sheet("Main", 0)
    list(map(lambda x: ws.cell(row=(column_1.index(x) + 1), column=1, value=x), column_1))
    list(map(lambda x: ws.cell(row=(column_2.index(x) + 1), column=2, value=x), column_2))
    list(map(lambda x: ws.cell(row=(column_3.index(x) + 1), column=3, value=x), column_3))
    list(map(lambda x: ws.cell(row=(column_4.index(x) + 1), column=4, value=x), column_4))
    list(map(lambda x: ws.cell(row=(column_5.index(x) + 1), column=5, value=x), column_5))
    list(map(lambda x: ws.cell(row=(column_6.index(x) + 1), column=6, value=x), column_6))

    # ADD EMAIL-DECODED

    for row in ws.iter_rows(min_row=1, max_col=1, max_row=100):
        for cell in row:
            if cell.value == 'Электронная почта:':
                ws.cell(row=cell.row, column=2, value=email_decoded)

    # ----------------------------

    file_name = 'Заказ' + '_' + str(order_number) + '.xlsx'
    wb.save(file_name)


if __name__ == '__main__':
    user_name = ''
    password_ = ''

    orders_list_out, s = login(user_name, password_)

    for order_out in [orders_list_out[0]]:
        letter_transform(order_out)
